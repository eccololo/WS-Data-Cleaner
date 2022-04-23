import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from termcolor import colored
import time
import sys
from art import logo

# ============= CONSTANTS ==============
# This is number of maximum columns that user can delete in data excel file.
# User can specify it in user mode.
MAX_COL_TO_DELETE = 30
MIN_COL_TO_DELETE = 1

# This is a list of unnecessary columns that Web Scraper creates and that we want to delete.
LIST_OF_UNNECESSARY_COLUMNS =  ["web-scraper-order", "web-scraper-start-url", "main-cat-link",
                                "main-cat-link-href", "pagination", "pagination-href", "item-link", "item-link-href"]

# ================ TESTS =================================

def check_if_rows_from_main_data_excel_files_equals_main_root_excel_file():
    """This function checkes if rows from all main files are equal to one root main
    excel file."""

    test_sum_of_rows_from_main_data_excel_files = 0
    test_of_rows_from_main_root_data_excel_file = 0

    test_current_file_path = os.getcwd()

    test_data_dir_file_path = os.path.join(test_current_file_path, "Data")
    
    if os.path.exists(test_data_dir_file_path):
        test_list_of_data_dirs = list(os.listdir(test_data_dir_file_path))
        
        # We go through all dirs which are in main Data dir.
        for test_data_dir in test_list_of_data_dirs:

            test_data_excel_files_location = os.path.join(test_data_dir_file_path, test_data_dir)
            # Taking list of all files in main dir.
            test_list_of_data_excel_files = list(os.listdir(test_data_excel_files_location))
            
            # We go through all excel files and dirs in category dir. There are excel file and
            # 'Done' and 'Main' dir.
            for test_data_excel_file_or_dir in test_list_of_data_excel_files:

                test_if_this_is_a_file_or_dir = os.path.join(test_data_dir_file_path, test_data_dir, test_data_excel_file_or_dir)

                if os.path.isdir(test_if_this_is_a_file_or_dir):
                    if test_data_excel_file_or_dir == "Main":
                        test_main_excel_file = list(os.listdir(test_if_this_is_a_file_or_dir))[0]
                        test_main_excel_file_path = os.path.join(test_if_this_is_a_file_or_dir, test_main_excel_file)
                        test_main_wb = load_workbook(test_main_excel_file_path)
                        test_main_ws = test_main_wb.active
                        test_sum_of_rows_from_main_data_excel_files += test_main_ws.max_row
                        # We deduct one row for header names row.
                        test_sum_of_rows_from_main_data_excel_files -= 1

    root_excel_file_path = os.path.join(test_current_file_path, "Products-All-Data.xlsx")

    if not os.path.isdir(root_excel_file_path):
        test_main_root_wb = load_workbook(root_excel_file_path)
        test_main_root_ws = test_main_root_wb.active
        test_of_rows_from_main_root_data_excel_file += test_main_root_ws.max_row

    # We add one because we add header names row.
    test_sum_of_rows_from_main_data_excel_files += 1
    print_result_of_test_data_excel_files_rows_equals_to_main_excel_file(test_sum_of_rows_from_main_data_excel_files, test_of_rows_from_main_root_data_excel_file, True)


def check_if_rows_from_data_excel_files_equals_main_excel_file():
    """This function checkes if number of rows from data excel files equals the 
    number in main excel files."""

    tests_msg = colored("[Tests:]", "green")
    explanation_msg = colored('[Explanation]:', 'green')
    results_msg = colored('[Results]:', 'green')
    email_msg = colored("mateusz.hyla.job@gmail.com", "red")

    print(tests_msg)
    print(f"{explanation_msg} Below you can compare how many rows there is in all data excel files,")
    print("in this category and how many rows there is in main excel output file.")
    print("If this two number are the same it means that program copied all data successfuly.")
    print("Green color means success, red means that something went wrong.")
    print(f"If you see red, you can contact admin for future explanation at: {email_msg}.")
    print(results_msg)

    test_sum_of_rows_from_data_excel_files = 0
    test_sum_of_rows_from_main_data_excel_files = 0

    test_current_file_path = os.getcwd()

    test_data_dir_file_path = os.path.join(test_current_file_path, "Data")
    
    if os.path.exists(test_data_dir_file_path):
        test_list_of_data_dirs = list(os.listdir(test_data_dir_file_path))
        
        # We go through all dirs which are in main Data dir.
        for test_data_dir in test_list_of_data_dirs:
            test_data_excel_files_location = os.path.join(test_data_dir_file_path, test_data_dir)
            # Taking list of all files in main dir.
            test_list_of_data_excel_files = list(os.listdir(test_data_excel_files_location))
            
            # We go through all excel files and dirs in category dir. There are excel file and
            # 'Done' and 'Main' dir.
            for test_data_excel_file_or_dir in test_list_of_data_excel_files:

                test_if_this_is_a_file_or_dir = os.path.join(test_data_dir_file_path, test_data_dir, test_data_excel_file_or_dir)

                if not os.path.isdir(test_if_this_is_a_file_or_dir):
                    test_wb = load_workbook(test_if_this_is_a_file_or_dir)
                    test_ws = test_wb.active
                    test_sum_of_rows_from_data_excel_files += test_ws.max_row
                    # We deduct row that is header names row.
                    test_sum_of_rows_from_data_excel_files -= 1
                else:
                    if test_data_excel_file_or_dir == "Main":
                        test_main_excel_file = list(os.listdir(test_if_this_is_a_file_or_dir))[0]
                        test_main_excel_file_path = os.path.join(test_if_this_is_a_file_or_dir, test_main_excel_file)
                        test_main_wb = load_workbook(test_main_excel_file_path)
                        test_main_ws = test_main_wb.active
                        test_sum_of_rows_from_main_data_excel_files = test_main_ws.max_row
            
            # We add one row to add header names row.
            test_sum_of_rows_from_data_excel_files += 1

            print_result_of_test_data_excel_files_rows_equals_to_main_excel_file(test_sum_of_rows_from_data_excel_files, test_sum_of_rows_from_main_data_excel_files)
            time.sleep(0.25)
            test_sum_of_rows_from_data_excel_files = 0


def print_result_of_test_data_excel_files_rows_equals_to_main_excel_file(rows_data_excel_files, rows_main_excel_file, root=False):
    """This function prints if the number of rows from data excel file is equal to rows from
    main excel file."""

    dots_green = colored('***********************************', 'green')
    dots_red = colored('***********************************', 'red')
    
    if root:
        root_msg = "Root"
        final_msg = colored("[Final Check]:", "green")
    else:
        root_msg = ""
        final_msg = ""

    is_equal = rows_data_excel_files == rows_main_excel_file
    if not is_equal:
        results_msg_bool = colored('False', 'red')
        not_equal_number_msg_1_excel_file = colored(rows_data_excel_files, "red")
        not_equal_number_msg_2_main_file = colored(rows_main_excel_file, "red")
    else:
        not_equal_number_msg_1_excel_file = colored(rows_data_excel_files, "green")
        not_equal_number_msg_2_main_file = colored(rows_main_excel_file, "green")
        results_msg_bool = colored('True', 'green')

    if is_equal:
        print(final_msg)
        print(dots_green)
        print(f"Data Excel Files: {not_equal_number_msg_1_excel_file}")
        print(f"Main {root_msg} Data Excel File: {not_equal_number_msg_2_main_file}")
        print(f"[Equal]: {results_msg_bool}")
        print(dots_green)
    else:
        print(final_msg)
        print(dots_red)
        print(f"Data Excel Files: {not_equal_number_msg_1_excel_file}")
        print(f"Main {root_msg} Data Excel File: {not_equal_number_msg_2_main_file}")
        print(f"[Equal]: {results_msg_bool}")
        print(dots_red)



# ================ END OF TESTS ==========================


# ========= FUNCTIONS ===============


def change_main_dir_names():
    """This function changes main dir names with excel data files to those withut any spaces.
    For example we have dir name 'Fire Stoves'. It will chage it to 'FireStoves'. I created
    this function because previously when dir had name with spaces I got FileNotFoundError
    during saving excel file."""
    change_partial_main_dir_names(" ", "")
    change_partial_main_dir_names("-", "")
    change_partial_main_dir_names("_", "")


def change_partial_main_dir_names(search_for, replace_with):
    """This function changes main dir names with excel that will be without unwanted chars
    like spaces, dashes or underscores."""

    path  = os.path.join(os.getcwd(), "Data")
    filenames = os.listdir(path)
    
    # We delete characters like space or _ or - in dir name. 
    for filename in filenames:
        if filename.find(search_for):
            os.rename(os.path.join(path, filename), os.path.join(path, filename.replace(search_for, replace_with)))


def print_welcome_message():
    """This function prints welcome message."""

    program = colored('Program:', 'green')
    author = colored('Author:', 'green')
    creation_date = colored('Creation Date:', 'green')
    contact = colored('Contact:', 'green')
    description = colored('Description:', 'green')
    dots = colored('***********************************', 'green')

    print(dots)
    print(f"{program} WS Data Cleaner - Automatyzacja Obróbki Danych Surowych Plików z Web Scrapera.")
    print(f"{author} Mateusz Hyla, Specjalista ds. E-Commerce.")
    print(f"{creation_date} 30.12.2021")
    print(f"{contact} mateusz.hyla.job@gmail.com")
    print(dots)
    print(f"{description}")
    print("Welcome to program that automates process of cleaning data from Web Scraper.")
    print("It deletes unnecessary column from raw excel data files,")
    print("and copy all that data from many files to main excel file.")
    print("So at the input you have raw excel files from Web Scraper.")
    print("And at the output you have cleaned excel files with data and")
    print("main excel file with all data combined.")
    print(dots)


def take_from_user_initial_data_auto_or_user():
    """This function takes from user initial data like auto or user and returns it."""
    initial_data = colored('Initial Data:', 'green')
    dots_green = colored('***********************************', 'green')
    dots_red = colored('***********************************', 'red')
    invalid_choice = colored("You can choose only 'user' or 'auto' option. Program terminated!", 'red')

    print(dots_green)
    print(initial_data)
    print("You can pass to this program some initial data which are for example:")
    print("1. Header column names in main excel file with all data of products.")
    print("2. Number of deleted unnecessary columns in raw web scraper data file starting from 1st to x.")
    print("But if you want program can automaticaly handle it for you.")
    print("Do you want to pass this data by yourself or leave the program to hanlde it for you?")
    user_choice_auto_or_user = input("(auto/user): ")
    print(dots_green)

    # If user typed invalid input we terminate program.
    if user_choice_auto_or_user.lower() not in ["auto", "user"]:
        print(dots_red)
        sys.exit(f"{invalid_choice}")
    
    return user_choice_auto_or_user


def take_from_user_initial_data_num_of_columns_to_delete():
    """This function takes from user number of coulmns he wants to delete and prints
    error message if this number is to big or to low or prints message that this columns are deleted.
    It also prints a hint message."""

    initial_data = colored('Initial Data:', 'green')
    dots_green = colored('***********************************', 'green')
    dots_red = colored('***********************************', 'red')
    invalid_choice = colored(f"You can enter only integer number between {MIN_COL_TO_DELETE} and {MAX_COL_TO_DELETE}. Program terminated.", 'red')
    contact_admin = colored(f"If you have more column to deleted than {MAX_COL_TO_DELETE} you can choose 'auto' option or contact admin (mateusz.hyla.job@gmail.com).", 'green')
    hint_msg = colored("Hint:", "green")

    print(dots_green)
    print(initial_data)
    print("Output excel files created by Web Scraper has a few additional")
    print("working columns that are not necessary to scraped data.")
    print("You can now specify how many columns do you want to delete starting from first column.")
    print("Columns can only be deleted in order from first one to specified one.")
    print(dots_green)
    print(hint_msg)
    print("If you don't know how many columns you need to delete you can open an excel file with scraped data")
    print("and check which column are to delete. You must remember that this column must be near each other.")
    print("You can only delete columns in continuous way for example from 1 to 5.")
    print("'User' option doesn't allow deleting column that are spread and separated.")
    print("If in you worksheet there are unnecessary column mixed with those you want to keep")
    print("You can use 'auto' option instead.")
    print(dots_green)

    # We check if the user entered a number. If not we terminate the program.
    try:
        print(dots_green)
        columns_num_to_delete = int(input("How many columns would you like to delete?: "))
        if columns_num_to_delete > MAX_COL_TO_DELETE or columns_num_to_delete < MIN_COL_TO_DELETE: 
            # Info message if a user would like to delete more columns than max number of delete.
            print(dots_green)
            print(contact_admin)
            print(dots_green)
            print(dots_red)
            sys.exit(f"{invalid_choice}")
        print(dots_green)
    except ValueError:
        print(dots_red)
        sys.exit(f"{invalid_choice}")

    for i in range(1, columns_num_to_delete + 1):
        time.sleep(0.25)
        print(f"Deleting {i} column ...")
        
    return columns_num_to_delete


def create_header_column_names(wb, h_names_list):
    """This function creates a header column names in a worksheet.
    We pass to it a list of header column names. It will write this 
    names to Excel first row and according columns."""
    ws = wb.active
    for indeks, h_name in enumerate(h_names_list):
        indeks = indeks + 1
        ws.cell(row=1, column=indeks).value = h_name


def save_data_excel_file(wb, file_path):
    """This function saves data excel file."""

    dots_red = colored('***********************************', 'red')
    error_file_msg_1 = colored("[Error-no.3]:", "red")
    error_file_msg_2 = colored("FilePath:", "red")

    try:
        wb.save(file_path)
    except FileNotFoundError:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving data excel file.")
        print("The program is calling FileNotFoundError.")
        print(f"{error_file_msg_2} {file_path}")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()
    except:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving data excel file.")
        print("The program is calling default exception.")
        print(f"{error_file_msg_2} {file_path}")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()


def save_main_excel_file(wb):
    """This function is creating a dir 'main' in dir where all data files are
    and it is saving in it main excel file."""

    dots_green = colored('***********************************', 'green')
    dots_red = colored('***********************************', 'red')
    error_file_msg_1 = colored("[Error-no.2]:", "red")

    main_dir_file_path = os.path.join(data_dir_file_path, data_dir, "Main")
    
    # If there is not yet created this dir we are creating it.
    if not os.path.isdir(main_dir_file_path):
        os.mkdir(main_dir_file_path)
    
    main_excel_file_save_location = os.path.join(main_dir_file_path, main_excel_file_name) 
    
    try:
        wb.save(filename=main_excel_file_save_location)
    except FileNotFoundError:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving main excel file.")
        print("The program is calling FileNotFoundError.")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()
    except:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving main excel file.")
        print("The program is calling default exception.")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()

    save_file_msg_1 = colored("[Success]", "red")
    save_file_msg_2 = colored(f"{main_excel_file_save_location}", "green")
    print(dots_green)
    print(save_file_msg_1, "Main excel file saved at: ", save_file_msg_2)
    print(dots_green)


def delete_unnecessary_columns_from_data_file(ws):
    """This function is deleting unneccesary columns from data exel file."""

    list_of_column_num_to_delete = []
    
    # We loop through only first row and all column and if we encounter unneccesary column
    # we add it to delete list.
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            if cell.value in LIST_OF_UNNECESSARY_COLUMNS:
                list_of_column_num_to_delete.append(cell.column)

    # Here we delete all columns that are unnecessary.
    for index, column_num_to_delete in enumerate(list_of_column_num_to_delete):
        if index > 0:
            deletion_number = index
            column_num_to_delete -= deletion_number
        ws.delete_cols(column_num_to_delete)


def get_header_column_names_from_data_excel_file(ws):
    """This function is taking header column names from data excel file and 
    returns it as a list."""
    list_of_header_column_names = []
    for col in ws.iter_cols(min_row=1, min_col=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            list_of_header_column_names.append(cell.value)

    return list_of_header_column_names


def get_data_from_data_excel_file(ws):
    """This function get all data from spreadsheet passed to it and returns
    every row as a list in a list of lists."""
    list_of_lists_of_data = []
    list_of_data_row = []
    for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in col:
            list_of_data_row.append(cell.value)
        list_of_lists_of_data.append(list_of_data_row)
        list_of_data_row = []

    return list_of_lists_of_data


def write_all_data_to_main_excel_file(wb, data_lists):
    """This function writes all data from data excel files to main excel file."""
    ws = wb.active
    
    row_num = 2
    column_num = 1

    for data_list in data_lists:
        for ws_cell_value in data_list:
            ws.cell(row=row_num, column=column_num).value = ws_cell_value
            column_num += 1
        column_num = 1
        row_num += 1 


def export_all_data_to_one_main_root_excel_file(data_lists, header_column_names):
    """This function write all data from main data excel files to one root main excel file."""
    
    dots_green = colored('***********************************', 'green')
    dots_red = colored('***********************************', 'red')
    error_file_msg_1 = colored("[Error-no.4]:", "red")
    main_excel_root_file_save_location = os.path.join(os.getcwd(), "Products-All-Data.xlsx")
    
    wb_root_main = Workbook()
    ws_root_main = wb_root_main.active

    # Creating header column names in main root excel file.
    create_header_column_names(wb_root_main, header_column_names)

    row_num = 2
    column_num = 1

    # Writing all data to main root excel file.
    for data_list in data_lists:
        for ws_cell_value in data_list:
            ws_root_main.cell(row=row_num, column=column_num).value = ws_cell_value
            column_num += 1
        column_num = 1
        row_num += 1 
    
    # Saving main root excel file.
    try:
        wb_root_main.save(main_excel_root_file_save_location)
    except FileNotFoundError:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving main root excel file.")
        print("The program is calling FileNotFoundError.")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()
    except:
        print(dots_red)
        print(f"{error_file_msg_1} Something is wrong with file path to a saving main root excel file.")
        print("The program is calling default exception.")
        print("To solve this you can contact admin at mateusz.hyla.job@gmail.com.")
        print(dots_red)
        sys.exit()

    # Success message of writing main root excel file.
    save_file_msg_1 = colored("[Success]", "red")
    save_file_msg_2 = colored(f"{main_excel_root_file_save_location}", "green")
    print(dots_green)
    print(save_file_msg_1, "Main root excel file saved at: ", save_file_msg_2)
    print(dots_green)


def do_create_custom_column_names_and_delete_some_column_in_user_more(user_choice):
    """This function do stuff that user specify in user mode in meaning it,
    delete some columns from excel files that user specify and it creates a custom
    header column names in main excel file that user specify."""

    # Below code is to create header column names in main Excel file.
    # User specify what names will be as header columns.
    header_column_names_in_main_excel_file = []
    counter = 1
    columns_num_to_delete = None
    if user_choice.lower() == "user":
        while True:
            h_name = input(f"Enter {counter}. header column name that will be in main Excel file (q - quit): ")
            if h_name.lower() == "q" or h_name.lower() == "quit":
                break
            header_column_names_in_main_excel_file.append(h_name)
            counter += 1
        columns_num_to_delete = take_from_user_initial_data_num_of_columns_to_delete()

    return columns_num_to_delete, header_column_names_in_main_excel_file

# ================ END OF FUNCTIONS ======================  


if __name__ == "__main__":

    # Changing names of the dirs with excel files.
    change_main_dir_names()

    dots_green = colored('***********************************', 'green')
    dots_red = colored('*************************************', 'red')

    print(dots_green)
    print(logo)

    # This function print welcome message
    print_welcome_message()
    # We set the mode of program. If it will work automatic or it will take initial data
    # from the user.
    user_choice_auto_or_user = take_from_user_initial_data_auto_or_user()

    current_file_path = os.getcwd()
    data_dir_file_path = os.path.join(current_file_path, "Data")

    if os.path.exists(data_dir_file_path):
        # We take a list of directories names from Data dir where 
        # are Excel files with datas. For example if in Data dir is
        # 'Plumbing' dir, where are stored files with that
        # category, we take that name.
        list_of_data_dirs = list(os.listdir(data_dir_file_path))

        # We create a list to store a all data from all main excel files.
        data_list_from_data_main_excel_files = []
        # We go through all dirs which are in main Data dir.

        columns_num_to_delete, header_column_names_in_main_excel_file = do_create_custom_column_names_and_delete_some_column_in_user_more(user_choice_auto_or_user)

        for data_dir in list_of_data_dirs:

            # Creating main Excel file name.
            main_excel_file_name = data_dir + "-data-all.xlsx"
            wb_main = Workbook()

            # We are creating header column names in main excel file that were specified
            # by user in user mode.
            create_header_column_names(wb_main, header_column_names_in_main_excel_file)

            # Extracting data from excel files with data and writing it to main excel file
            data_excel_files_location = os.path.join(data_dir_file_path, data_dir)
            # Taking list of all files in main dir.
            list_of_data_excel_files = list(os.listdir(data_excel_files_location))
            list_of_header_column_names_in_main_excel_file = []
            data_list_from_data_excel_file = []
            only_once_flag = True
            for data_excel_file in list_of_data_excel_files:
                # We check if this is a proper excel file we want.
                path_to_excel_data_file = os.path.join(data_excel_files_location, data_excel_file)
                if os.path.isfile(path_to_excel_data_file):
                    # We open this excel file to read.
                    wb_data = load_workbook(path_to_excel_data_file)
                    ws_data = wb_data.active

                    if user_choice_auto_or_user.lower() == "user": 
                        # We are cleaning unnecessary columns.
                        ws_data.delete_cols(1, columns_num_to_delete)
                    else:
                        delete_unnecessary_columns_from_data_file(ws_data)

                    # Here we extract header column names from data file to write it to main excel file.
                    if user_choice_auto_or_user.lower() == "auto" and only_once_flag:
                        header_column_names_in_main_excel_file = get_header_column_names_from_data_excel_file(ws_data)
                        only_once_flag = False

                    # Here we create a list that contains a lists of every row in data excel files
                    # we go through in this loop.
                    data_list_from_data_excel_file.extend(get_data_from_data_excel_file(ws_data))

                    # If there is not yet created Done dir we are creating it.
                    done_dir_file_path = os.path.join(data_excel_files_location, "Done")
                    if not os.path.isdir(done_dir_file_path):
                        os.mkdir(done_dir_file_path)

                    path_to_save_excel_data_file = os.path.join(data_excel_files_location, "Done", data_excel_file)
                    save_data_excel_file(wb_data, path_to_save_excel_data_file)

            # Here we create header column names in main excel file.
            create_header_column_names(wb_main, header_column_names_in_main_excel_file)

            # We create a list of lists of all rows in all main data excel files.
            data_list_from_data_main_excel_files.extend(data_list_from_data_excel_file)

            # Here we write all data from data excel files to main excel file.
            write_all_data_to_main_excel_file(wb_main, data_list_from_data_excel_file)
            save_main_excel_file(wb_main)
        
        # We are exporting all data from main files to one main root file.
        export_all_data_to_one_main_root_excel_file(data_list_from_data_main_excel_files, header_column_names_in_main_excel_file)

    else:
        # Error when there is no 'Data' folder in root dir.
        print(dots_red)
        error_1_msg = colored("[Error-no.1]: There is no 'Data' dir in main folder.", "red")
        print(error_1_msg) 
        print(dots_red)
        print(dots_green)
        explanation_msg = colored("All folders with data files should be in main folder called 'Data'. \nContact admin at mateusz.hyla.job@gmail.com for support.", "green")
        print(explanation_msg)
        print(dots_green)
        sys.exit()
    
    check_if_rows_from_data_excel_files_equals_main_excel_file()
    check_if_rows_from_main_data_excel_files_equals_main_root_excel_file()

 